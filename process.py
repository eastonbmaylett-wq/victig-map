#!/usr/bin/env python3
"""
Victig TAT Map — Data Processor (v2)
Supports:
  - Old CSV format: Jurisdiction, Search TAT, Start Date/Time, Search Name
  - New Excel format: Jurisdiction, Exact TAT, Start Date/Time, Completion Date/Time,
                      Unique Jurisdiction, Avg. TAT, Max. TAT, Min. TAT, >2 weeks TAT, <0.007 TAT
Usage: python3 process.py <path-to-csv-or-xlsx>
"""
import csv, json, io, re, sys
from datetime import datetime
from collections import defaultdict
from pathlib import Path

BASE     = Path(__file__).parent
IN_PATH  = sys.argv[1] if len(sys.argv) > 1 else str(BASE / "uploaded.csv")
IN_FILE  = Path(IN_PATH)

STATE_FIPS = {
    '01':'AL','02':'AK','04':'AZ','05':'AR','06':'CA','08':'CO','09':'CT',
    '10':'DE','11':'DC','12':'FL','13':'GA','15':'HI','16':'ID','17':'IL',
    '18':'IN','19':'IA','20':'KS','21':'KY','22':'LA','23':'ME','24':'MD',
    '25':'MA','26':'MI','27':'MN','28':'MS','29':'MO','30':'MT','31':'NE',
    '32':'NV','33':'NH','34':'NJ','35':'NM','36':'NY','37':'NC','38':'ND',
    '39':'OH','40':'OK','41':'OR','42':'PA','44':'RI','45':'SC','46':'SD',
    '47':'TN','48':'TX','49':'UT','50':'VT','51':'VA','53':'WA','54':'WV',
    '55':'WI','56':'WY'
}
COLOR_STATUS = {'#dea922':'delay','#de7612':'high_tat','#cf4027':'significant'}

# ── Helpers ───────────────────────────────────────────────────────────────
def classify(s):
    s = s.lower()
    if 'county' in s and 'criminal' in s: return 'county_criminal'
    if 'state' in s and ('criminal' in s or 'court' in s or 'search' in s): return 'statewide'
    return 'other'

def compute_stats(vals):
    if not vals: return None
    return {'avg': round(sum(vals)/len(vals), 6),
            'max': round(max(vals), 6),
            'min': round(min(vals), 6),
            'count': len(vals)}

def parse_dt(s):
    if not s: return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%m/%d/%Y %H:%M', '%Y-%m-%dT%H:%M:%S',
                '%m/%d/%Y %H:%M:%S', '%Y-%m-%d %H:%M'):
        try: return datetime.strptime(str(s).strip(), fmt)
        except: pass
    return None

def safe_float(v):
    try: return float(v)
    except: return None

# ── Load file ─────────────────────────────────────────────────────────────
suffix = IN_FILE.suffix.lower()
raw_rows = []

if suffix in ('.xlsx', '.xlsm'):
    import openpyxl
    wb = openpyxl.load_workbook(str(IN_FILE), read_only=True, data_only=True)
    # prefer the main data sheet
    sheet_names = wb.sheetnames
    main_sheet = None
    for name in sheet_names:
        n = name.lower()
        if 'turnaround' in n or 'search' in n or 'tat' in n:
            main_sheet = wb[name]; break
    if not main_sheet:
        main_sheet = wb.active
    for row in main_sheet.iter_rows(values_only=True):
        raw_rows.append(['' if v is None else str(v) for v in row])
else:
    with open(str(IN_FILE), newline='', encoding='utf-8-sig') as f:
        lines = f.readlines()
    # skip non-header preamble rows (before the actual header)
    start = 0
    for i, line in enumerate(lines):
        if 'jurisdiction' in line.lower() or 'search tat' in line.lower():
            start = i; break
    reader = csv.reader(io.StringIO(''.join(lines[start:])))
    for row in reader:
        raw_rows.append(row)

if not raw_rows:
    print("ERROR: No rows found in file"); sys.exit(1)

# ── Detect columns ────────────────────────────────────────────────────────
header = raw_rows[0]
hmap   = {h.strip().lower(): i for i, h in enumerate(header)}

def col(name): return hmap.get(name.lower())

# Required columns (try both old + new names)
COL_JUR      = col('jurisdiction')
COL_EXACT    = col('exact tat')
COL_SEARCH   = col('search tat')
COL_TAT      = COL_EXACT if COL_EXACT is not None else COL_SEARCH   # prefer exact
COL_START    = col('start date/time')
COL_COMPLETE = col('completion date/time')
COL_SNAME    = col('search name')
COL_CLIENT   = col('client name')
COL_CAND     = col('candidate name')
COL_FILE     = col('file #')

# New pre-computed aggregate columns
COL_UJUR     = col('unique jurisdiction')
COL_AVG      = col('avg. tat')
COL_MAX      = col('max. tat')
COL_MIN      = col('min. tat')
COL_GT2W     = col('>2 weeks tat')
COL_LT007    = col('<0.007 tat')

is_new_format = COL_EXACT is not None

# Allow raw format: no TAT column but start + completion present → compute on the fly
COMPUTE_TAT = COL_TAT is None and COL_START is not None and COL_COMPLETE is not None

if COL_JUR is None:
    print(f"ERROR: No 'Jurisdiction' column found. Header: {header}"); sys.exit(1)
if COL_TAT is None and not COMPUTE_TAT:
    print(f"ERROR: No TAT column found and no Start+Completion columns to compute from. Header: {header}"); sys.exit(1)

if COMPUTE_TAT:
    print("Format: RAW (TAT computed from Start/Completion dates)")
else:
    print(f"Format: {'NEW (Exact TAT)' if is_new_format else 'OLD (Search TAT)'}")
print(f"Columns: {len(header)} | Rows: {len(raw_rows)-1}")

# ── Parse rows ────────────────────────────────────────────────────────────
rows, latest = [], {}
# Pre-computed jurisdiction stats from new format
precomp = {}  # jur -> {avg, max, min, gt2w, lt007}

for raw in raw_rows[1:]:
    def get(c): return raw[c].strip() if c is not None and c < len(raw) else ''

    jur      = get(COL_JUR)
    date_str = get(COL_START)
    search   = get(COL_SNAME)

    # Compute TAT from start/completion if no TAT column
    if COMPUTE_TAT:
        comp_dt  = parse_dt(get(COL_COMPLETE))
        start_dt = parse_dt(date_str)
        if start_dt and comp_dt and comp_dt > start_dt:
            tat_str = str((comp_dt - start_dt).total_seconds() / 86400)
        else:
            tat_str = ''
    else:
        tat_str = get(COL_TAT)

    # Collect pre-computed stats (new format only)
    if COL_UJUR is not None:
        ujur = get(COL_UJUR)
        if ujur:
            avg = safe_float(get(COL_AVG))
            mx  = safe_float(get(COL_MAX))
            mn  = safe_float(get(COL_MIN))
            gt2 = safe_float(get(COL_GT2W))
            lt7 = safe_float(get(COL_LT007))
            if avg is not None:
                precomp[ujur] = {'avg': avg, 'max': mx, 'min': mn,
                                 'gt2weeks': int(gt2) if gt2 else 0,
                                 'lt10min':  int(lt7) if lt7 else 0}

    if not jur or not tat_str or not date_str: continue
    tat = safe_float(tat_str)
    dt  = parse_dt(date_str)
    if tat is None or dt is None or tat < 0: continue

    rows.append({'jur': jur, 'tat': tat, 'dt': dt, 'type': classify(search)})

    # Latest search per jurisdiction
    if jur not in latest or dt > latest[jur]['dt']:
        comp_str = get(COL_COMPLETE) if COL_COMPLETE else ''
        latest[jur] = {
            'dt': dt, 'tat': tat, 'search': search,
            'client':     get(COL_CLIENT),
            'candidate':  get(COL_CAND) if COL_CAND is not None else '',
            'completion': comp_str,
            'file_num':   get(COL_FILE),
        }

if not rows:
    print("ERROR: No valid data rows parsed"); sys.exit(1)

max_date = max(r['dt'] for r in rows)

# ── Period stats (computed from individual rows) ──────────────────────────
jur_periods = defaultdict(lambda: defaultdict(list))
jur_types   = defaultdict(lambda: defaultdict(list))

for r in rows:
    age = (max_date - r['dt']).days
    if age > 180: continue
    for label, days in [('30d',30),('60d',60),('90d',90),('180d',180)]:
        if age <= days:
            jur_periods[r['jur']][label].append(r['tat'])
    jur_types[r['jur']][r['type']].append(r['tat'])

# Trend (30d vs 30-60d)
jur_trend = {}
for r in rows:
    age = (max_date - r['dt']).days
    if age > 60: continue
    jur = r['jur']
    if jur not in jur_trend: jur_trend[jur] = {'recent':[],'prior':[]}
    if age <= 30: jur_trend[jur]['recent'].append(r['tat'])
    else:         jur_trend[jur]['prior'].append(r['tat'])

period_stats = {}
for jur, pdata in jur_periods.items():
    period_stats[jur] = {p: compute_stats(v) for p,v in pdata.items()}

# Rank within each period
for period in ['30d','60d','90d','180d']:
    ranked = sorted(
        [(j, s[period]['avg']) for j,s in period_stats.items() if s.get(period)],
        key=lambda x: x[1]
    )
    for rank,(jur,_) in enumerate(ranked, 1):
        period_stats[jur][period]['rank']  = rank
        period_stats[jur][period]['total'] = len(ranked)

type_stats = {j:{t:compute_stats(v) for t,v in d.items()} for j,d in jur_types.items()}

trends = {}
for jur, d in jur_trend.items():
    if d['recent'] and d['prior']:
        diff = sum(d['prior'])/len(d['prior']) - sum(d['recent'])/len(d['recent'])
        trends[jur] = 0 if abs(diff)<0.1 else (1 if diff>0 else -1)

latest_out = {}
for j, v in latest.items():
    latest_out[j] = {
        'tat':        v['tat'],
        'date':       v['dt'].strftime('%b %d, %Y %H:%M'),
        'search':     v['search'],
        'completion': v['completion'],
        'client':     v['client'],
        'candidate':  v.get('candidate',''),
        'file_num':   v['file_num'],
    }

# ── Parse Outliers sheet (if Excel with multiple sheets) ─────────────────
# Maps jurisdiction -> list of {tat, reason, search, client}
outliers_by_jur = defaultdict(list)
if suffix in ('.xlsx', '.xlsm'):
    try:
        import openpyxl as _opx
        _wb = _opx.load_workbook(str(IN_FILE), read_only=True, data_only=True)
        _ws = None
        for _sn in _wb.sheetnames:
            if 'outlier' in _sn.lower(): _ws = _wb[_sn]; break
        if _ws:
            _rows = list(_ws.iter_rows(values_only=True))
            _hdr  = {str(h).strip().lower(): i for i,h in enumerate(_rows[0]) if h}
            def _g(row, key):
                i = _hdr.get(key)
                return str(row[i]).strip() if i is not None and i < len(row) and row[i] else ''
            for r in _rows[1:]:
                jur    = _g(r,'jurisdiction')
                reason = _g(r,'reason')
                tat    = safe_float(_g(r,'search tat'))
                if jur and reason:
                    outliers_by_jur[jur].append({
                        'tat':    tat,
                        'reason': reason,
                        'search': _g(r,'search name'),
                        'client': _g(r,'client name'),
                    })
    except Exception as e:
        print(f'Outliers sheet skipped: {e}')

# ── Load SimpleMaps base (names, existing status/descriptions) ───────────
sm_file = BASE / "simplemaps-base.json"
with open(sm_file) as f:
    map_data = json.load(f)

county_db = {}
for fips, info in map_data.items():
    abbrev = STATE_FIPS.get(fips[:2], '')
    county_db[fips] = {
        'fips':        fips,
        'name':        info.get('name',''),
        'state':       abbrev,
        'status':      COLOR_STATUS.get(info.get('color',''), 'ok'),
        'description': info.get('description',''),
        'periods':  None,
        'types':    None,
        'trend':    0,
        'latest':   None,
        'precomp':  None,
        'outliers': None,
    }

# Add US territories from counties-10m.json
TERRITORY_FIPS = {'60':'AS','66':'GU','69':'MP','72':'PR','78':'VI'}
topo_file = BASE / "counties-10m.json"
if topo_file.exists():
    with open(topo_file) as f:
        topo_data = json.load(f)
    for geom in topo_data['objects']['counties']['geometries']:
        fips = str(geom['id'])
        if fips[:2] in TERRITORY_FIPS and fips not in county_db:
            name = geom.get('properties', {}).get('name', fips)
            county_db[fips] = {
                'fips': fips, 'name': name,
                'state': TERRITORY_FIPS[fips[:2]],
                'status': 'ok', 'description': '',
                'periods': None, 'types': None, 'trend': 0,
                'latest': None, 'precomp': None, 'outliers': None,
            }

# ── Statewide averages ────────────────────────────────────────────────────
# Group ALL rows by state prefix (e.g. 'MT-Lincoln County' → 'MT')
# and compute average TAT across every search in that state.
state_all_tats   = defaultdict(list)   # all rows, all time
state_period_tats = defaultdict(lambda: defaultdict(list))

for r in rows:
    jur = r['jur']
    if '-' not in jur: continue
    st = jur.split('-', 1)[0].strip().upper()
    if len(st) != 2: continue          # must be a 2-letter state code
    state_all_tats[st].append(r['tat'])
    age = (max_date - r['dt']).days
    for label, days in [('30d',30),('60d',60),('90d',90),('180d',180)]:
        if age <= days:
            state_period_tats[st][label].append(r['tat'])

state_avg = {}
for st, vals in state_all_tats.items():
    periods = {p: compute_stats(v) for p, v in state_period_tats[st].items()}
    state_avg[st] = {
        'avg':     round(sum(vals) / len(vals), 6),
        'count':   len(vals),
        'periods': periods,
    }

# ── Match jurisdictions to FIPS ───────────────────────────────────────────
def norm(s): return re.sub(r'[^a-z]','',s.lower())

sm_map = {}
for fips, c in county_db.items():
    state, name = c['state'], c['name'].upper()
    for jur in set(list(period_stats.keys()) + list(precomp.keys())):
        if '-' not in jur: continue
        parts = jur.split('-', 1)
        if parts[0].strip().upper() != state: continue
        county = parts[1].strip().upper()
        for suffix in ['', ' COUNTY', ' PARISH', ' BOROUGH', ' CITY']:
            if norm(re.sub(suffix+'$', '', county).strip()) == norm(name):
                sm_map[fips] = jur; break

for fips, jur in sm_map.items():
    county_db[fips]['periods']  = period_stats.get(jur)
    county_db[fips]['types']    = type_stats.get(jur)
    county_db[fips]['trend']    = trends.get(jur, 0)
    county_db[fips]['latest']   = latest_out.get(jur)
    county_db[fips]['precomp']  = precomp.get(jur)
    county_db[fips]['outliers'] = outliers_by_jur.get(jur) or None

# Attach state avg to every county (whether matched or not)
for fips, c in county_db.items():
    st = c.get('state', '')
    c['state_avg'] = state_avg.get(st)  # None if no data for that state

# ── Preserve admin overrides ──────────────────────────────────────────────
existing_file = BASE / "county-data.json"
if existing_file.exists():
    with open(existing_file) as f:
        existing = json.load(f)
    for fips, c in existing.get('counties', {}).items():
        if fips in county_db and c.get('_admin_override'):
            county_db[fips]['status']          = c['status']
            county_db[fips]['description']      = c['description']
            county_db[fips]['_admin_override']  = True

# ── Write output ──────────────────────────────────────────────────────────
out = {
    'maxDate':     max_date.strftime('%Y-%m-%d'),
    'isNewFormat': is_new_format,
    'stateAvg':    state_avg,   # top-level for easy frontend lookup
    'counties':    county_db,
}
with open(BASE / "county-data.json", 'w') as f:
    json.dump(out, f)

matched_states = len(state_avg)
print(f"Done. {len(sm_map)} counties matched | {len(set(list(period_stats.keys())+list(precomp.keys())))} unique jurisdictions | {matched_states} states with avg data. Max date: {max_date.date()}")
