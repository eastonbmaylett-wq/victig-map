from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
import subprocess, shutil, os, hashlib, json, csv, io, sqlite3, datetime, asyncio
import urllib.request as _urllib_req
from pathlib import Path

app = FastAPI(docs_url=None, redoc_url=None, openapi_url=None)  # disable API docs

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["GET", "POST"],
    allow_headers=["Content-Type"],
)

BASE = Path(__file__).parent
DATA_DIR    = Path(os.environ.get("DATA_DIR", "/data" if Path("/data").exists() or os.environ.get("RAILWAY_ENVIRONMENT") else str(BASE)))
try:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    DATA_DIR = BASE  # fallback to app directory if /data isn't writable
DATA_FILE   = DATA_DIR / "county-data.json"
CONFIG_FILE = DATA_DIR / "site-config.json"

# Seed volume from bundled files on first boot
_BUNDLED_DATA   = BASE / "county-data.json"
_BUNDLED_CONFIG = BASE / "site-config.json"
if not DATA_FILE.exists() and _BUNDLED_DATA.exists():
    try: shutil.copy(_BUNDLED_DATA, DATA_FILE)
    except Exception: pass
if not CONFIG_FILE.exists() and _BUNDLED_CONFIG.exists():
    try: shutil.copy(_BUNDLED_CONFIG, CONFIG_FILE)
    except Exception: pass

from fastapi.staticfiles import StaticFiles
app.mount("/static", StaticFiles(directory=BASE / "static"), name="static")

# ── Analytics DB ────────────────────────────────────────────────────────────
ANALYTICS_DB = DATA_DIR / "analytics.db"

def _analytics_conn():
    con = sqlite3.connect(str(ANALYTICS_DB))
    con.execute("""CREATE TABLE IF NOT EXISTS events (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        ts        TEXT NOT NULL,
        ip        TEXT,
        ua        TEXT,
        event     TEXT NOT NULL,
        county_id TEXT,
        county    TEXT,
        state     TEXT
    )""")
    con.commit()
    return con

def _geoip_sync(ip: str) -> str:
    """Return 'City, State' for an IP using ip-api.com (called in a thread)."""
    try:
        # Skip private / loopback IPs
        if ip.startswith(('10.','172.','192.168.','127.','::1','fd')):
            return ''
        url = f"http://ip-api.com/json/{ip}?fields=status,city,regionName"
        with _urllib_req.urlopen(url, timeout=3) as r:
            d = json.loads(r.read())
            if d.get('status') == 'success':
                parts = [d.get('city',''), d.get('regionName','')]
                return ', '.join(p for p in parts if p)
    except Exception:
        pass
    return ''

def _log_event(request: Request, event: str, county_id=None, county=None, state=None):
    try:
        ip = request.headers.get('x-forwarded-for', request.client.host).split(',')[0].strip()
        ua = request.headers.get('user-agent', '')[:200]
        ts = datetime.datetime.utcnow().isoformat()
        con = _analytics_conn()
        con.execute("INSERT INTO events (ts,ip,ua,event,county_id,county,state) VALUES (?,?,?,?,?,?,?)",
                    (ts, ip, ua, event, county_id, county, state))
        con.commit()
        row_id = con.execute('SELECT last_insert_rowid()').fetchone()[0]
        con.close()
        # Fire-and-forget geo lookup for visit/embed events with no state set
        if event in ('visit', 'embed') and not state:
            async def _enrich():
                try:
                    loop = asyncio.get_event_loop()
                    loc = await loop.run_in_executor(None, _geoip_sync, ip)
                    if loc:
                        c = _analytics_conn()
                        c.execute('UPDATE events SET state=? WHERE id=?', (loc, row_id))
                        c.commit(); c.close()
                except Exception:
                    pass
            asyncio.ensure_future(_enrich())
    except Exception:
        pass  # never break the map over analytics

# ── Cache-bust version: hash of all static files combined
import hashlib as _hashlib
try:
    _static_hash = _hashlib.md5(
        b''.join(f.read_bytes() for f in sorted((BASE / 'static').glob('*')) if f.is_file())
    ).hexdigest()[:8]
except Exception:
    _static_hash = 'dev'

# Seed data files from repo if not yet on volume
try:
    for _f in ["county-data.json", "site-config.json"]:
        _src = BASE / _f
        _dst = DATA_DIR / _f
        if _src.exists() and not _dst.exists():
            import shutil as _shutil
            _shutil.copy2(_src, _dst)
except Exception:
    pass
PW_HASH   = "b3121997c76507dc7adcf3ca13ee60d519cbc3c72a176527e8ba575fc13f3406"

# ── Security headers ──────────────────────────────────────────────────────
SECURITY_HEADERS = {
    "X-Content-Type-Options":    "nosniff",
    "X-XSS-Protection":          "1; mode=block",
    "Referrer-Policy":           "no-referrer",
    "Permissions-Policy":        "geolocation=(), camera=(), microphone=()",
    # X-Frame-Options intentionally omitted - map is designed to be embedded in TazWorks/InstaScreen
}

def secure(response: Response):
    for k, v in SECURITY_HEADERS.items():
        response.headers[k] = v
    return response

# ── Auth ──────────────────────────────────────────────────────────────────
def check_auth(password: str):
    if not password:
        raise HTTPException(status_code=401, detail="Password required")
    h = hashlib.sha256(password.encode()).hexdigest()
    if h != PW_HASH:
        raise HTTPException(status_code=401, detail="Invalid password")

# ── Block any direct file access not explicitly allowed ───────────────────
BLOCKED_EXTENSIONS = {".py", ".json", ".csv", ".txt", ".toml", ".env", ".sh"}
BLOCKED_FILES      = {"requirements.txt", "Procfile", "uploaded.csv",
                      "simplemaps-base.json", "process.py", "main.py"}

@app.middleware("http")
async def block_raw_files(request: Request, call_next):
    path = request.url.path.lstrip("/")
    # Block direct access to sensitive files
    if path in BLOCKED_FILES:
        return JSONResponse(status_code=404, content={"detail": "Not found"})
    # Block any file extension that shouldn't be public
    suffix = Path(path).suffix.lower()
    ALLOWED_FILES = {"county-data.json", "counties-10m.json", "d3.min.js", "topojson-client.min.js", "site-config.json"}
    if suffix in BLOCKED_EXTENSIONS and path not in ALLOWED_FILES:
        return JSONResponse(status_code=404, content={"detail": "Not found"})
    response = await call_next(request)
    for k, v in SECURITY_HEADERS.items():
        response.headers[k] = v
    return response

# ── Public routes ─────────────────────────────────────────────────────────
def _inject_version(html_path: Path) -> Response:
    """Serve HTML with ?v=hash appended to all /static/ asset URLs."""
    import re as _re
    html = _re.sub(r'(/static/[^"]+)', lambda m: m.group(1) + f'?v={_static_hash}', html_path.read_text())
    return Response(content=html, media_type="text/html",
                    headers={"Cache-Control": "no-store, no-cache, must-revalidate"})

@app.get("/")
def root(request: Request):
    _log_event(request, 'visit')
    return _inject_version(BASE / "index.html")

@app.get("/county-data.json")
def get_data():
    """Serve only the sanitized county stats - no PII."""
    return FileResponse(
        DATA_FILE,
        media_type="application/json",
        headers={"Cache-Control": "no-cache"}
    )

@app.get("/counties-10m.json")
def get_topo():
    return FileResponse(
        BASE / "counties-10m.json",
        media_type="application/json",
        headers={"Cache-Control": "public, max-age=86400"}
    )

@app.get("/d3.min.js")
def get_d3():
    return FileResponse(
        BASE / "d3.min.js",
        media_type="application/javascript",
        headers={"Cache-Control": "public, max-age=86400"}
    )

@app.get("/topojson-client.min.js")
def get_topojson():
    return FileResponse(
        BASE / "topojson-client.min.js",
        media_type="application/javascript",
        headers={"Cache-Control": "public, max-age=86400"}
    )


# ── Admin routes (password required for all writes) ───────────────────────
@app.get("/site-config.json")
def get_site_config():
    return FileResponse(CONFIG_FILE, media_type="application/json",
                        headers={"Cache-Control": "no-cache"})

@app.post("/admin/site-config")
async def update_site_config(payload: dict):
    check_auth(payload.get("password", ""))
    config = payload.get("config", {})
    allowed_keys = {"title", "subtitle", "statusLabels", "legendNotes",
                    "badgeLabel", "fastestTitle", "slowestTitle", "brandText"}
    safe = {k: v for k, v in config.items() if k in allowed_keys}
    with open(CONFIG_FILE, "w") as f:
        json.dump(safe, f, indent=2)
    return {"ok": True}

@app.get("/admin")
def admin_page():
    return _inject_version(BASE / "admin.html")

@app.get("/embed")
def embed_page(request: Request):
    _log_event(request, 'embed')
    return _inject_version(BASE / "embed.html")


@app.get("/victig-logo-white.png")
def logo_white():
    return FileResponse(BASE / "victig-logo-white.png", media_type="image/png")

@app.get("/victig-logo.png")
def logo_orig():
    return FileResponse(BASE / "victig-logo.png", media_type="image/png")

@app.post("/admin/update-county")
async def update_county(payload: dict):
    check_auth(payload.get("password", ""))
    fips        = payload.get("fips")
    status      = payload.get("status", "ok")
    description = payload.get("description", "")
    if not fips:
        raise HTTPException(status_code=400, detail="fips required")
    if status not in ("ok", "delay", "high_tat", "significant", "closed"):
        raise HTTPException(status_code=400, detail="invalid status")
    with open(DATA_FILE) as f:
        data = json.load(f)
    if fips not in data["counties"]:
        raise HTTPException(status_code=404, detail="County not found")
    data["counties"][fips]["status"]          = status
    data["counties"][fips]["description"]     = description
    data["counties"][fips]["_admin_override"] = True
    # Store manual stat overrides (strip empty strings)
    overrides = payload.get("overrides", {})
    clean_ov = {k: v for k, v in overrides.items() if v}
    if clean_ov:
        data["counties"][fips]["_overrides"] = clean_ov
    elif "_overrides" in data["counties"][fips]:
        del data["counties"][fips]["_overrides"]
    with open(DATA_FILE, "w") as f:
        json.dump(data, f)
    return {"ok": True}

@app.post("/admin/upload")
async def upload_csv(password: str, file: UploadFile = File(...)):
    check_auth(password)
    name = (file.filename or "").lower()
    raw = await file.read()
    tmp = Path("/tmp/victig_upload.csv")

    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            # Pick the main data sheet (prefer sheet with most rows / matching name)
            ws = None
            for sname in wb.sheetnames:
                n = sname.lower()
                if 'turnaround' in n or ('search' in n and 'tat' not in n) or \
                   ('tat' in n and 'fastest' not in n and 'outlier' not in n):
                    ws = wb[sname]; break
            if ws is None:
                # Fall back to largest sheet by row count
                best, best_rows = None, 0
                for sname in wb.sheetnames:
                    s = wb[sname]
                    r = s.max_row or 0
                    if r > best_rows:
                        best_rows, best = r, sname
                ws = wb[best] if best else wb.active
            rows = list(ws.iter_rows(values_only=True))
            with open(tmp, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in rows:
                    writer.writerow(["" if v is None else str(v) for v in row])
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Excel parse error: {e}")
    elif name.endswith(".xls"):
        raise HTTPException(status_code=400, detail="Old .xls format not supported - please Save As .xlsx in Excel first")
    elif name.endswith(".tsv") or name.endswith(".txt"):
        # Tab-separated → convert to comma CSV
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text), delimiter="\t")
        with open(tmp, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in reader:
                writer.writerow(row)
    else:
        # Treat as CSV (covers .csv and any unknown format)
        with open(tmp, "wb") as f:
            f.write(raw)

    result = subprocess.run(
        ["python3", str(BASE / "process.py"), str(tmp)],
        capture_output=True, text=True, cwd=str(BASE)
    )
    tmp.unlink(missing_ok=True)
    if result.returncode != 0:
        err = (result.stderr or result.stdout or "Unknown error")[:800]
        raise HTTPException(status_code=500, detail=err)
    return {"ok": True, "message": result.stdout.strip() or "Map data updated successfully"}

@app.post("/admin/preview-columns")
async def preview_columns(password: str, file: UploadFile = File(...)):
    """Return the column headers + first 3 rows of an uploaded file without processing it."""
    check_auth(password)
    name = file.filename or ""
    raw = await file.read()
    rows = []
    try:
        if name.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = None
            for sname in wb.sheetnames:
                n = sname.lower()
                if 'turnaround' in n or ('search' in n and 'tat' not in n) or \
                   ('tat' in n and 'fastest' not in n and 'outlier' not in n):
                    ws = wb[sname]; break
            if ws is None:
                best, best_rows = None, 0
                for sname in wb.sheetnames:
                    s = wb[sname]
                    r = s.max_row or 0
                    if r > best_rows:
                        best_rows, best = r, sname
                ws = wb[best] if best else wb.active
            rows = [list(r) for r in list(ws.iter_rows(values_only=True))[:6]]
        else:
            text = raw.decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = [r for _, r in zip(range(6), reader)]
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"rows": rows}

# ── Doc description parser ────────────────────────────────────────────────
STATE_NAMES_FULL = {
    'ALABAMA','ALASKA','ARIZONA','ARKANSAS','CALIFORNIA','COLORADO','CONNECTICUT',
    'DELAWARE','FLORIDA','GEORGIA','HAWAII','IDAHO','ILLINOIS','INDIANA','IOWA',
    'KANSAS','KENTUCKY','LOUISIANA','MAINE','MARYLAND','MASSACHUSETTS','MICHIGAN',
    'MINNESOTA','MISSISSIPPI','MISSOURI','MONTANA','NEBRASKA','NEVADA',
    'NEW HAMPSHIRE','NEW JERSEY','NEW MEXICO','NEW YORK','NORTH CAROLINA',
    'NORTH DAKOTA','OHIO','OKLAHOMA','OREGON','PENNSYLVANIA','RHODE ISLAND',
    'SOUTH CAROLINA','SOUTH DAKOTA','TENNESSEE','TEXAS','UTAH','VERMONT',
    'VIRGINIA','WASHINGTON','WEST VIRGINIA','WISCONSIN','WYOMING',
    'DISTRICT OF COLUMBIA', 'HAWAII',
}
STATE_ABBREV = {
    'ALABAMA':'AL','ALASKA':'AK','ARIZONA':'AZ','ARKANSAS':'AR','CALIFORNIA':'CA',
    'COLORADO':'CO','CONNECTICUT':'CT','DELAWARE':'DE','FLORIDA':'FL','GEORGIA':'GA',
    'HAWAII':'HI','IDAHO':'ID','ILLINOIS':'IL','INDIANA':'IN','IOWA':'IA',
    'KANSAS':'KS','KENTUCKY':'KY','LOUISIANA':'LA','MAINE':'ME','MARYLAND':'MD',
    'MASSACHUSETTS':'MA','MICHIGAN':'MI','MINNESOTA':'MN','MISSISSIPPI':'MS',
    'MISSOURI':'MO','MONTANA':'MT','NEBRASKA':'NE','NEVADA':'NV',
    'NEW HAMPSHIRE':'NH','NEW JERSEY':'NJ','NEW MEXICO':'NM','NEW YORK':'NY',
    'NORTH CAROLINA':'NC','NORTH DAKOTA':'ND','OHIO':'OH','OKLAHOMA':'OK',
    'OREGON':'OR','PENNSYLVANIA':'PA','RHODE ISLAND':'RI','SOUTH CAROLINA':'SC',
    'SOUTH DAKOTA':'SD','TENNESSEE':'TN','TEXAS':'TX','UTAH':'UT','VERMONT':'VT',
    'VIRGINIA':'VA','WASHINGTON':'WA','WEST VIRGINIA':'WV','WISCONSIN':'WI',
    'WYOMING':'WY'
}
SKIP_PARAS = {
    'NEW UPDATES','ONGOING/GENERAL COURT DELAYS','DEAR CLIENT PARTNERS,',
    'SINCERELY,','PLEASE LET US KNOW','IT\'S AN HONOR','THANK YOU',
}

def parse_desc_docx(raw_bytes):
    """Parse Victig court delays Word doc -> list of (state_abbrev, county_name_or_None, description)
    Stores text word-for-word. Handles:
    - Inline single county:   'MACON COUNTY is closed due to...'
    - Inline multi-county:    'COCONINO, NAVAJO, PINAL and YAVAPAI COUNTIES are...'
    - Standalone sub-headers: 'LOS ANGELES' followed by description paragraphs
    - State-level notes:      paragraphs with no county match
    - 'ALL [STATE] COUNTIES': treated as state-level
    """
    import zipfile, xml.etree.ElementTree as ET
    with zipfile.ZipFile(io.BytesIO(raw_bytes)) as z:
        xml_data = z.read('word/document.xml').decode('utf-8')
    root = ET.fromstring(xml_data)
    W = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
    paras = []
    for p in root.iter(f'{{{W}}}p'):
        text = ''.join(r.text or '' for r in p.iter(f'{{{W}}}t')).strip()
        if text: paras.append(text)

    results = []                # (state_abbrev, county_or_None, description)
    current_state  = None
    current_county = None       # set by standalone sub-headers like 'LOS ANGELES'
    county_paras   = []         # description paragraphs collected under a sub-header county

    def flush_county():
        """Emit any collected sub-header county paragraphs as a single entry."""
        nonlocal current_county, county_paras
        if county_paras and current_state:
            state_ab = STATE_ABBREV.get(current_state)
            if state_ab:
                results.append((state_ab, current_county, ' '.join(county_paras)))
        current_county = None
        county_paras   = []

    COUNTY_RE  = re.compile(r'\b([A-Z][A-Z\s]+?)\s+COUNTY\b')
    COUNTIES_RE = re.compile(r'\bCOUNTIES\b')

    for para in paras:
        upper = para.upper().strip()

        # ── Skip boilerplate ──────────────────────────────────────────────
        # Normalize smart quotes so curly apostrophes match SKIP_PARAS strings
        upper_norm = upper.replace('\u2019', "'").replace('\u2018', "'")
        if upper_norm in SKIP_PARAS: continue
        if any(skip in upper_norm for skip in SKIP_PARAS): continue

        # ── State header ─────────────────────────────────────────────────
        if upper in STATE_NAMES_FULL:
            flush_county()
            current_state = upper
            continue

        if not current_state: continue
        state_ab = STATE_ABBREV.get(current_state)
        if not state_ab: continue

        # ── Standalone all-caps sub-header: 'LOS ANGELES' ────────────────
        # Must check BEFORE the length filter — headers like 'LOS ANGELES' are short
        if len(para) <= 50 and re.match(r'^[A-Z][A-Z\s]+$', para):
            flush_county()
            county = re.sub(r'\s+COUNTY$', '', upper).strip()
            current_county = county
            continue

        # Skip very short lines with no substance (after sub-header check)
        if len(para) < 20: continue

        # ── Multi-county sentence: '... COUNTIES ...' ─────────────────────
        if COUNTIES_RE.search(para):
            flush_county()
            m = re.search(r'([\w\s,]+?)\s+COUNTIES\b', para)
            if m:
                raw_list = m.group(1).strip()
                # Normalize ', and' → ',' before splitting so 'LUCAS, and MEDINA' works
                raw_list = re.sub(r',\s+and\s+', ', ', raw_list, flags=re.I)
                parts = re.split(r',\s*|\s+and\s+', raw_list, flags=re.I)
                counties = [
                    p.strip().upper() for p in parts
                    if p.strip()
                    and 2 <= len(p.strip()) <= 20          # real county names are short
                    and re.match(r'^[A-Z\s]+$', p.strip().upper())  # no digits or symbols
                    and p.strip().upper() not in STATE_NAMES_FULL
                    and not p.strip().upper().startswith('ALL')
                ]
                if counties:
                    for c in counties:
                        results.append((state_ab, c, para))
                    continue
            # Fallback: treat as state-level (e.g. 'ALL SOUTH CAROLINA COUNTIES')
            results.append((state_ab, None, para))
            continue

        # ── Single COUNTY inline: 'MACON COUNTY is closed...' ────────────
        single = COUNTY_RE.findall(para)
        if single:
            flush_county()
            for c in single:
                c = c.strip()
                if 2 <= len(c) <= 40:
                    results.append((state_ab, c, para))
            continue

        # ── Regular description paragraph ─────────────────────────────────
        if len(para) >= 30:
            if current_county is not None:
                county_paras.append(para)   # collect under sub-header county
            else:
                results.append((state_ab, None, para))  # state-level note

    flush_county()  # emit any trailing sub-header county
    return results

@app.post("/admin/upload-desc-doc")
async def upload_desc_doc(password: str, file: UploadFile = File(...)):
    """Import a Word .docx descriptions document and update county descriptions."""
    check_auth(password)
    raw = await file.read()
    name = (file.filename or '').lower()
    if not name.endswith('.docx'):
        raise HTTPException(status_code=400, detail='Upload a .docx Word document')

    try:
        entries = parse_desc_docx(raw)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'Parse error: {e}')

    with open(DATA_FILE) as f:
        data = json.load(f)
    counties = data['counties']

    # Build reverse lookup: norm(state+name) -> fips
    def norm(s): return re.sub(r'[^a-z]', '', s.lower())
    lookup = {}
    for fips, c in counties.items():
        key = norm((c.get('state') or '') + (c.get('name') or ''))
        lookup[key] = fips
        lookup[norm(c.get('name') or '')] = fips

    # Clear ALL existing descriptions on fresh doc import
    # (new doc is authoritative — wipes manual entries too)
    cleared = 0
    for fips, c in counties.items():
        if c.get('description') or c.get('_desc_doc'):
            c['description'] = ''
            c.pop('_desc_doc', None)
            cleared += 1

    # Apply new descriptions
    updated, skipped = 0, []
    for state_ab, county_name, desc in entries:
        fips = None
        if county_name:
            key = norm(state_ab + county_name)
            fips = lookup.get(key) or lookup.get(norm(county_name))
        else:
            # State-level: apply to all counties in state without a specific desc
            for fp, c in counties.items():
                if c.get('state') == state_ab and not c.get('description'):
                    c['description'] = desc
                    c['_desc_doc'] = True
                    updated += 1
            continue
        if fips and fips in counties:
            counties[fips]['description'] = desc
            counties[fips]['_desc_doc'] = True
            # Auto-detect status from description keywords
            low = desc.lower()
            if any(kw in low for kw in ('closed indefinitely','closed for','courthouse is closed',
                                         'inaccessible','no access','shut down','closed due',
                                         'operations ceased','closed until','closed effective',
                                         'not yet resumed','have not resumed')):
                counties[fips]['status'] = 'closed'
            elif counties[fips].get('status') == 'ok':
                counties[fips]['status'] = 'delay'
            updated += 1
        else:
            skipped.append(f'{state_ab}-{county_name}')

    with open(DATA_FILE, 'w') as f:
        json.dump(data, f)

    msg = f'Updated {updated} counties, cleared {cleared} old entries.'
    if skipped:
        msg += f' Unmatched ({len(skipped)}): {", ".join(skipped[:8])}'
    return {'ok': True, 'updated': updated, 'cleared': cleared,
            'skipped': len(skipped), 'message': msg}

@app.post("/admin/upload-descriptions")
async def upload_descriptions(password: str, file: UploadFile = File(...)):
    """Import a CSV/Excel with jurisdiction + description columns."""
    check_auth(password)
    name = (file.filename or "").lower()
    raw  = await file.read()
    rows = []

    try:
        if name.endswith(".xlsx") or name.endswith(".xlsm"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows = [["" if v is None else str(v).strip() for v in r]
                    for r in ws.iter_rows(values_only=True)]
        else:
            text = raw.decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = [[c.strip() for c in r] for r in reader]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Parse error: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="No rows found")

    # Auto-detect columns
    header = [h.lower() for h in rows[0]]
    def find_col(*keywords):
        for kw in keywords:
            for i, h in enumerate(header):
                if kw in h: return i
        return None

    jur_col  = find_col('jurisdiction','county','location','state','name')
    desc_col = find_col('description','note','delay','text','reason','message','comment')
    status_col = find_col('status','level','severity')

    if jur_col is None or desc_col is None:
        raise HTTPException(status_code=400,
            detail=f"Could not find jurisdiction+description columns. Headers: {rows[0]}")

    # Load county data
    with open(DATA_FILE) as f:
        data = json.load(f)
    counties = data["counties"]

    # Build a lookup: jurisdiction string -> list of fips
    import re
    STATE_ABBREVS = {
        'AL','AK','AZ','AR','CA','CO','CT','DE','DC','FL','GA','HI','ID','IL','IN',
        'IA','KS','KY','LA','ME','MD','MA','MI','MN','MS','MO','MT','NE','NV','NH',
        'NJ','NM','NY','NC','ND','OH','OK','OR','PA','RI','SC','SD','TN','TX','UT',
        'VT','VA','WA','WV','WI','WY'
    }
    def norm(s): return re.sub(r'[^a-z]','',s.lower())

    # Build reverse lookup: norm(state+name) -> fips
    lookup = {}
    for fips, c in counties.items():
        key = norm(c.get('state','') + c.get('name',''))
        lookup[key] = fips
        # Also try just name
        lookup[norm(c.get('name',''))] = fips

    updated = 0
    skipped = []
    for row in rows[1:]:
        if len(row) <= max(jur_col, desc_col): continue
        jur  = row[jur_col].strip()
        desc = row[desc_col].strip()
        status_val = row[status_col].strip().lower() if status_col and status_col < len(row) else None
        if not jur or not desc: continue

        # Try matching: STATE-COUNTY, or just COUNTY, or full name
        fips = None
        # Try STATE-COUNTY format
        if '-' in jur:
            parts = jur.split('-', 1)
            state, county = parts[0].strip().upper(), parts[1].strip()
            key = norm(state + county)
            fips = lookup.get(key)
        if not fips:
            fips = lookup.get(norm(jur))
        if not fips:
            # Try matching just the county name part
            for k, v in lookup.items():
                if norm(jur) in k or k in norm(jur):
                    fips = v; break

        if fips and fips in counties:
            counties[fips]['description'] = desc
            counties[fips]['_admin_override'] = True
            if status_val and status_val in ('ok','delay','high_tat','significant'):
                counties[fips]['status'] = status_val
            updated += 1
        else:
            skipped.append(jur)

    with open(DATA_FILE, 'w') as f:
        json.dump(data, f)

    msg = f"Updated {updated} counties."
    if skipped:
        msg += f" Skipped {len(skipped)} unmatched: {', '.join(skipped[:5])}"
        if len(skipped) > 5: msg += f" (+{len(skipped)-5} more)"
    return {"ok": True, "updated": updated, "skipped": len(skipped), "message": msg}

@app.post("/admin/update-state")
async def update_state(payload: dict):
    check_auth(payload.get("password", ""))
    state       = payload.get("state", "").upper()
    status      = payload.get("status", "ok")
    description = payload.get("description", "")
    if not state:
        raise HTTPException(status_code=400, detail="state required")
    if status not in ("ok", "delay", "high_tat", "significant", "closed"):
        raise HTTPException(status_code=400, detail="invalid status")
    with open(DATA_FILE) as f:
        data = json.load(f)
    updated = 0
    for fips, c in data["counties"].items():
        if c.get("state", "").upper() == state:
            c["status"]          = status
            c["description"]     = description
            c["_admin_override"] = True
            updated += 1
    with open(DATA_FILE, "w") as f:
        json.dump(data, f)
    return {"ok": True, "updated": updated}

import time as _time
_START_TIME = str(int(_time.time()))  # changes every container restart / deploy

@app.get("/api/version")
def version():
    return {"version": _START_TIME, "has_analytics": True}

@app.post("/api/log/click")
async def log_click(request: Request):
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    _log_event(request, 'click',
               county_id=payload.get('fips'),
               county=payload.get('county'),
               state=payload.get('state'))
    return {"ok": True}

@app.get("/api/analytics")
async def get_analytics(password: str):
    check_auth(password)
    con = _analytics_conn()
    rows = con.execute("""
        SELECT ts, ip, ua, event, county_id, county, state
        FROM events ORDER BY ts DESC LIMIT 2000
    """).fetchall()
    con.close()
    return [{"ts":r[0],"ip":r[1],"ua":r[2],"event":r[3],
             "county_id":r[4],"county":r[5],"state":r[6]} for r in rows]

@app.get("/api/analytics/summary")
async def get_analytics_summary(password: str, exclude_ip: str = None):
    check_auth(password)
    con = _analytics_conn()
    ex = exclude_ip or ''
    total_visits   = con.execute("SELECT COUNT(*) FROM events WHERE event='visit' AND (? = '' OR ip != ?)", (ex, ex)).fetchone()[0]
    unique_ips     = con.execute("SELECT COUNT(DISTINCT ip) FROM events WHERE event='visit' AND (? = '' OR ip != ?)", (ex, ex)).fetchone()[0]
    today_visits   = con.execute("SELECT COUNT(*) FROM events WHERE event='visit' AND ts >= date('now') AND (? = '' OR ip != ?)", (ex, ex)).fetchone()[0]
    top_counties   = con.execute("""
        SELECT county, state, COUNT(*) as cnt FROM events
        WHERE event='county_click' AND county IS NOT NULL AND (? = '' OR ip != ?)
        GROUP BY county, state ORDER BY cnt DESC LIMIT 10
    """, (ex, ex)).fetchall()
    recent         = con.execute("""
        SELECT ts, ip, ua, event, county, state FROM events
        ORDER BY ts DESC LIMIT 200
    """).fetchall()
    visitors       = con.execute("""
        SELECT ip,
               MIN(ts) as first_seen,
               MAX(ts) as last_seen,
               SUM(CASE WHEN event IN ('visit','embed') THEN 1 ELSE 0 END) as visits,
               GROUP_CONCAT(DISTINCT county) as counties,
               MAX(ua) as ua,
               MAX(state) as location
        FROM events
        WHERE ip IS NOT NULL AND (? = '' OR ip != ?)
        GROUP BY ip
        ORDER BY last_seen DESC
        LIMIT 200
    """, (ex, ex)).fetchall()
    by_day         = con.execute("""
        SELECT substr(ts,1,10) as day, COUNT(*) as cnt
        FROM events WHERE event='visit' AND (? = '' OR ip != ?)
        GROUP BY day ORDER BY day DESC LIMIT 30
    """, (ex, ex)).fetchall()
    top_states = con.execute("""
        SELECT state, COUNT(*) as cnt FROM events
        WHERE event='visit' AND state IS NOT NULL AND state != '' AND (? = '' OR ip != ?)
        GROUP BY state ORDER BY cnt DESC LIMIT 10
    """, (ex, ex)).fetchall()
    con.close()
    return {
        "total_visits": total_visits,
        "unique_ips": unique_ips,
        "today_visits": today_visits,
        "top_counties": [{"county":r[0],"state":r[1],"clicks":r[2]} for r in top_counties],
        "top_states": [{"state":r[0],"visits":r[1]} for r in top_states],
        "recent": [{"ts":r[0],"ip":r[1],"ua":r[2],"event":r[3],"county":r[4],"state":r[5]} for r in recent],
        "by_day": [{"day":r[0],"visits":r[1]} for r in by_day],
        "visitors": [{"ip":r[0],"first_seen":r[1],"last_seen":r[2],"visits":r[3],
                       "counties":r[4],"ua":r[5],"location":r[6]} for r in visitors],
    }

@app.post("/api/analytics/backfill-geo")
async def backfill_geo(password: str):
    """Enrich existing visit rows that have no state with IP geolocation."""
    check_auth(password)
    con = _analytics_conn()
    rows = con.execute(
        "SELECT id, ip FROM events WHERE event IN ('visit','embed') AND (state IS NULL OR state='') AND ip IS NOT NULL"
    ).fetchall()
    con.close()
    enriched = 0
    loop = asyncio.get_event_loop()
    for row_id, ip in rows:
        loc = await loop.run_in_executor(None, _geoip_sync, ip)
        if loc:
            c = _analytics_conn()
            c.execute('UPDATE events SET state=? WHERE id=?', (loc, row_id))
            c.commit(); c.close()
            enriched += 1
        await asyncio.sleep(0.05)  # stay under ip-api.com rate limit (45/min)
    return {"enriched": enriched, "total": len(rows)}

@app.get("/api/my-ip")
async def my_ip(request: Request):
    ip = request.headers.get('x-forwarded-for', request.client.host).split(',')[0].strip()
    return {"ip": ip}

@app.get("/analytics")
async def analytics_page():
    return FileResponse(BASE / "analytics.html")

@app.get("/test")
def test_page():
    html = """<!DOCTYPE html><html><head><title>Map Diagnostics</title></head><body>
<h2>Map Load Diagnostics</h2>
<div id='log' style='font-family:monospace;font-size:14px'></div>
<script>
const log = (msg, ok=true) => {
  const el = document.createElement('div');
  el.style.color = ok ? 'green' : 'red';
  el.textContent = (ok ? '✅ ' : '❌ ') + msg;
  document.getElementById('log').appendChild(el);
};
async function run(){
  log('Starting tests...', true);
  try {
    const r1 = await fetch('/d3.min.js');
    log(`d3.min.js: HTTP ${r1.status}, ${r1.headers.get('content-type')}, ${(await r1.arrayBuffer()).byteLength} bytes`);
  } catch(e){ log('d3.min.js FAILED: '+e.message, false); }
  try {
    const r2 = await fetch('/topojson-client.min.js');
    log(`topojson.min.js: HTTP ${r2.status}, ${(await r2.arrayBuffer()).byteLength} bytes`);
  } catch(e){ log('topojson FAILED: '+e.message, false); }
  try {
    const r3 = await fetch('/counties-10m.json');
    log(`counties-10m.json: HTTP ${r3.status}, ${(await r3.arrayBuffer()).byteLength} bytes`);
  } catch(e){ log('counties-10m.json FAILED: '+e.message, false); }
  try {
    const r4 = await fetch('/county-data.json');
    log(`county-data.json: HTTP ${r4.status}, ${(await r4.arrayBuffer()).byteLength} bytes`);
  } catch(e){ log('county-data.json FAILED: '+e.message, false); }
  log('All tests complete.');
}
run();
</script></body></html>"""
    return Response(content=html, media_type="text/html")

